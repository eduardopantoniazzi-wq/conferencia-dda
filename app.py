import streamlit as st
import pandas as pd
import pdfplumber
import re
import io
import unicodedata
from rapidfuzz.distance import JaroWinkler
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

# ─── Config ───────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Conferência DDA", page_icon="🔍", layout="wide")
LIMIAR_NOME = 0.75  # pode ser ajustado pelo usuário na sidebar

# ─── Helpers ──────────────────────────────────────────────────────────────────

def norm(s: str) -> str:
    if not s:
        return ""
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.lower()
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def similaridade(a: str, b: str) -> float:
    na, nb = norm(a), norm(b)
    if not na or not nb:
        return 0.0
    return JaroWinkler.normalized_similarity(na, nb)


def tol_valor(v: float) -> float:
    return max(v * 0.01, 1.00)


def norm_nf(s) -> str:
    if not s:
        return ""
    limpo = re.sub(r"[^0-9]", "", str(s))
    return limpo.lstrip("0") or ""


def norm_seu_num(s) -> str:
    """Remove non-digits, leading zeros and the last digit (parcel indicator)."""
    if not s:
        return ""
    limpo = re.sub(r"[^0-9]", "", str(s))
    sem_zeros = limpo.lstrip("0") or ""
    return sem_zeros[:-1] if len(sem_zeros) > 1 else sem_zeros


def parse_valor(s) -> float | None:
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s).strip()
    s = re.sub(r"[R$\s]", "", s)
    # 1.234,56 → 1234.56
    if re.search(r"\d\.\d{3},", s):
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def fmt_brl(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "—"
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_data(d) -> str:
    if d is None or (isinstance(d, float) and pd.isna(d)):
        return "—"
    if isinstance(d, (date,)):
        return d.strftime("%d/%m/%Y")
    if isinstance(d, str) and re.match(r"\d{4}-\d{2}-\d{2}", d):
        y, m, dd = d[:4], d[5:7], d[8:10]
        return f"{dd}/{m}/{y}"
    return str(d)


# ─── Leitores ──────────────────────────────────────────────────────────────────

def _parse_data_str(ds: str) -> str | None:
    ds = str(ds).strip()
    m = re.match(r"(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{4})", ds)
    if m:
        a, b2, y = int(m.group(1)), int(m.group(2)), m.group(3)
        if a > 12:
            return f"{y}-{str(b2).zfill(2)}-{str(a).zfill(2)}"
        return f"{y}-{str(a).zfill(2)}-{str(b2).zfill(2)}"
    m2 = re.match(r"(\d{4})[\/\-](\d{2})[\/\-](\d{2})", ds)
    if m2:
        return ds[:10]
    return None


def _detectar_colunas_por_conteudo(df_data: pd.DataFrame):
    """Detecta colunas por padrão de conteúdo quando os cabeçalhos não batem."""
    re_val  = re.compile(r"^\s*R?\$?\s*[\d.]+,\d{2}\s*$")
    re_data = re.compile(r"\d{2}[\/\-]\d{2}[\/\-]\d{4}|\d{4}[\/\-]\d{2}[\/\-]\d{2}")
    re_num  = re.compile(r"^\s*\d{5,}\s*$")

    sample = df_data.head(20)
    n_cols = len(df_data.columns)

    score_val  = []
    score_data = []
    score_nome = []
    score_snum = []

    for col in range(n_cols):
        vals = sample.iloc[:, col].astype(str)
        sv = sum(1 for v in vals if re_val.search(v) and v.strip() not in ("nan",""))
        sd = sum(1 for v in vals if re_data.search(v) and v.strip() not in ("nan",""))
        sn = sum(1 for v in vals if len(v.strip()) > 8 and not re_val.search(v)
                 and not re_data.search(v) and not re_num.match(v)
                 and v.strip() not in ("nan","") and not v.strip().isdigit())
        ss = sum(1 for v in vals if re_num.match(v.strip()))
        score_val.append(sv); score_data.append(sd)
        score_nome.append(sn); score_snum.append(ss)

    ci = int(pd.Series(score_data).idxmax()) if max(score_data) > 0 else -1
    vi = int(pd.Series(score_val).idxmax())  if max(score_val)  > 0 else -1
    # para nome: maior score mas que não seja data nem valor nem seu número
    excluir = {ci, vi}
    nome_scores = [(j, s) for j, s in enumerate(score_nome) if j not in excluir]
    bi = max(nome_scores, key=lambda x: x[1])[0] if nome_scores else -1
    snum_scores = [(j, s) for j, s in enumerate(score_snum) if j not in excluir and j != bi]
    si = max(snum_scores, key=lambda x: x[1])[0] if snum_scores and max(s for _,s in snum_scores) > 0 else -1

    return ci, bi, vi, si


def ler_dda_excel(uploaded) -> tuple[pd.DataFrame, dict]:
    df_raw = pd.read_excel(uploaded, header=None, dtype=str, engine="openpyxl")
    K_DATA  = ["data venc", "vencimento", "dt venc", "venc", "data"]
    K_BENEF = ["beneficiario original", "beneficiario", "favorecido", "nome", "sacado", "pagador"]
    K_VAL   = ["valor (r$)", "valor r$", "valor"]
    K_SNUM  = ["seu numero", "seu num", "seu n", "numero doc", "num doc", "documento"]

    header_row = None
    ci = bi = vi = si = -1
    header_cols = []

    for i, row in df_raw.iterrows():
        r = [norm(str(c)) for c in row]
        di  = next((j for j, c in enumerate(r) if any(k in c for k in K_DATA)),  -1)
        bii = next((j for j, c in enumerate(r) if any(k in c for k in K_BENEF)), -1)
        vii = next((j for j, c in enumerate(r) if any(k in c for k in K_VAL)),   -1)
        sii = next((j for j, c in enumerate(r) if any(k in c for k in K_SNUM)),  -1)
        if di >= 0 and bii >= 0 and vii >= 0:
            header_row = i; ci = di; bi = bii; vi = vii; si = sii
            header_cols = list(row)
            break

    info = {}
    if header_row is None:
        # fallback: detecta por conteúdo
        ci2, bi2, vi2, si2 = _detectar_colunas_por_conteudo(df_raw)
        ci, bi, vi, si = ci2, bi2, vi2, si2
        header_row = 0
        info["aviso"] = "Cabeçalho de DDA não reconhecido — colunas detectadas pelo conteúdo"
        info["colunas_brutas"] = list(df_raw.iloc[0])
    else:
        info["colunas_brutas"] = header_cols

    info["col_data"] = ci; info["col_benef"] = bi
    info["col_valor"] = vi; info["col_snum"] = si

    regs = []
    for _, row in df_raw.iloc[header_row + 1:].iterrows():
        vals = list(row)
        benef   = str(vals[bi]).strip() if 0 <= bi < len(vals) else ""
        valor   = parse_valor(vals[vi]) if 0 <= vi < len(vals) else None
        data_raw = str(vals[ci]).strip() if 0 <= ci < len(vals) else ""
        seu_num  = norm_seu_num(vals[si]) if 0 <= si < len(vals) else None

        if not benef or benef in ("nan", "") or valor is None or valor <= 0:
            continue

        data = _parse_data_str(data_raw) if data_raw and data_raw != "nan" else None
        regs.append({"data": data, "beneficiario": benef, "valor": valor, "seuNum": seu_num})

    return pd.DataFrame(regs), info


def ler_dda_pdf(uploaded) -> pd.DataFrame:
    regs = []
    re_val  = re.compile(r"(?:R\$\s*)?([\d]{1,3}(?:[.\s]?\d{3})*,\d{2})")
    re_data = re.compile(r"(\d{2}[\/\-]\d{2}[\/\-]\d{4})")
    re_num  = re.compile(r"^\d{6,}$")

    with pdfplumber.open(uploaded) as pdf:
        for page in pdf.pages:
            # Tenta extrair tabela estruturada primeiro
            tables = page.extract_tables()
            for table in tables:
                if not table or len(table) < 2:
                    continue
                # detecta índices de coluna no cabeçalho
                header = [norm(str(c)) for c in table[0]]
                K_D = ["venc", "data"]; K_B = ["benef", "favorec", "nome"]; K_V = ["valor"]; K_S = ["seu"]
                ci = next((j for j, h in enumerate(header) if any(k in h for k in K_D)), -1)
                bi = next((j for j, h in enumerate(header) if any(k in h for k in K_B)), -1)
                vi = next((j for j, h in enumerate(header) if any(k in h for k in K_V)), -1)
                si = next((j for j, h in enumerate(header) if any(k in h for k in K_S)), -1)

                if bi < 0 or vi < 0:
                    continue

                for row in table[1:]:
                    if not row:
                        continue
                    benef = str(row[bi]).strip() if bi < len(row) else ""
                    val_raw = str(row[vi]).strip() if vi < len(row) else ""
                    data_raw = str(row[ci]).strip() if ci >= 0 and ci < len(row) else ""
                    seu_raw  = str(row[si]).strip() if si >= 0 and si < len(row) else ""

                    vm = re_val.search(val_raw)
                    if not vm:
                        continue
                    valor = float(vm.group(1).replace(".", "").replace(" ", "").replace(",", "."))
                    if valor <= 0 or not benef or benef == "None":
                        continue

                    data = None
                    dm = re_data.search(data_raw)
                    if dm:
                        ds = dm.group(1)
                        a, b2, y = ds[:2], ds[3:5], ds[6:]
                        a_i, b_i = int(a), int(b2)
                        if a_i > 12:
                            data = f"{y}-{b2}-{a}"
                        else:
                            data = f"{y}-{a}-{b2}"

                    regs.append({"data": data, "beneficiario": benef,
                                 "valor": valor, "seuNum": norm_seu_num(seu_raw) if seu_raw else None})

            # Se tabela estruturada não deu resultado, usa extração por palavras
            if not regs:
                words = page.extract_words(x_tolerance=3, y_tolerance=3)
                if not words:
                    continue

                # agrupa por linha (y0)
                linhas: dict[int, list] = {}
                for w in words:
                    y_key = round(w["top"] / 4) * 4
                    linhas.setdefault(y_key, []).append(w)
                for y_key in linhas:
                    linhas[y_key].sort(key=lambda w: w["x0"])

                # detecta cabeçalho
                K_DATA  = ["venc", "data"]; K_BENEF = ["benef", "favorec"]; K_VAL = ["valor"]; K_SNUM = ["seu"]
                cD = cB = cV = cS = cab_y = None

                sorted_ys = sorted(linhas.keys())
                for y_key in sorted_ys:
                    row_words = linhas[y_key]
                    texts = [norm(w["text"]) for w in row_words]
                    full  = " ".join(texts)
                    if (any(k in full for k in K_DATA) and
                        any(k in full for k in K_BENEF) and
                        any(k in full for k in K_VAL)):
                        # posições X das colunas
                        for w, t in zip(row_words, texts):
                            if any(k in t for k in K_DATA)  and cD is None: cD = w["x0"]
                            if any(k in t for k in K_BENEF) and cB is None: cB = w["x0"]
                            if any(k in t for k in K_VAL)   and cV is None: cV = w["x0"]
                            if any(k in t for k in K_SNUM)  and cS is None: cS = w["x0"]
                        cab_y = y_key
                        break

                TX = 60
                if cab_y is not None:
                    for y_key in sorted_ys:
                        if y_key <= cab_y:
                            continue
                        row_words = linhas[y_key]
                        get = lambda cx: next((w["text"] for w in row_words if abs(w["x0"] - cx) <= TX), None) if cx is not None else None
                        pB = get(cB); pV = get(cV)
                        if not pB or not pV:
                            continue
                        vm = re_val.search(str(pV))
                        if not vm:
                            continue
                        valor = float(vm.group(1).replace(".", "").replace(" ", "").replace(",", "."))
                        if valor <= 0:
                            continue
                        pD = get(cD); pS = get(cS)
                        dm = re_data.search(str(pD)) if pD else None
                        data = None
                        if dm:
                            ds = dm.group(1)
                            a, b2, y = ds[:2], ds[3:5], ds[6:]
                            data = f"{y}-{(b2 if int(a)<=12 else a)}-{(a if int(a)<=12 else b2)}"
                        regs.append({"data": data, "beneficiario": pB.strip(),
                                     "valor": valor, "seuNum": norm_seu_num(pS) if pS else None})
                else:
                    # fallback puro por linha
                    for y_key in sorted_ys:
                        row_words = linhas[y_key]
                        texts = [w["text"] for w in row_words]
                        full  = " ".join(texts)
                        vm = re_val.search(full)
                        if not vm:
                            continue
                        valor = float(vm.group(1).replace(".", "").replace(" ", "").replace(",", "."))
                        if valor <= 0:
                            continue
                        dm = re_data.search(full)
                        data = None
                        if dm:
                            ds = dm.group(1)
                            a, b2, y = ds[:2], ds[3:5], ds[6:]
                            data = f"{y}-{(b2 if int(a)<=12 else a)}-{(a if int(a)<=12 else b2)}"
                        # beneficiário: texto mais longo que não seja data nem valor
                        cands = [t for t in texts
                                 if not re_val.search(t) and not re_data.search(t) and len(t) > 3]
                        benef = max(cands, key=len) if cands else ""
                        if len(benef) < 3:
                            continue
                        sn_cell = next((t for t in texts if re_num.match(t.replace(" ", ""))), None)
                        regs.append({"data": data, "beneficiario": benef.strip(),
                                     "valor": valor, "seuNum": norm_seu_num(sn_cell) if sn_cell else None})

    return pd.DataFrame(regs) if regs else pd.DataFrame(columns=["data", "beneficiario", "valor", "seuNum"])


def ler_sistema(uploaded) -> pd.DataFrame:
    df_raw = pd.read_excel(uploaded, header=None, dtype=str, engine="openpyxl")
    K_DATA  = ["vencimento", "venc", "data"]
    K_NOME  = ["terceiro", "nome", "fornecedor"]
    K_VAL   = ["vlr. nom", "vlr nom", "valor nominal", "valor"]
    K_NF    = ["nota fis", "nota fiscal", "nf", "num nota", "numero nota"]

    header_row = ci = ni = vi = nfi = -1

    for i, row in df_raw.iterrows():
        r = [norm(str(c)) for c in row]
        di   = next((j for j, c in enumerate(r) if any(k in c for k in K_DATA)), -1)
        nii  = next((j for j, c in enumerate(r) if any(k in c for k in K_NOME)), -1)
        vii  = next((j for j, c in enumerate(r) if any(k in c for k in K_VAL)),  -1)
        nfii = next((j for j, c in enumerate(r) if any(k in c for k in K_NF)),   -1)
        if di >= 0 and nii >= 0 and vii >= 0:
            header_row = i; ci = di; ni = nii; vi = vii; nfi = nfii
            break

    if header_row < 0:
        st.warning("Sistema: cabeçalho não detectado — assumindo colunas 0,1,2")
        header_row = 0; ci = 0; ni = 1; vi = 2; nfi = -1

    regs = []
    for _, row in df_raw.iloc[header_row + 1:].iterrows():
        vals = list(row)
        nome  = str(vals[ni]).strip() if ni < len(vals) else ""
        valor = parse_valor(vals[vi]) if vi < len(vals) else None
        nf_raw = str(vals[nfi]).strip() if nfi >= 0 and nfi < len(vals) else ""

        if not nome or nome in ("nan", "") or valor is None or valor <= 0:
            continue

        data_raw = vals[ci] if ci < len(vals) else None
        data = None
        if data_raw and str(data_raw) not in ("nan", ""):
            ds = str(data_raw).strip()
            m = re.match(r"(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{4})", ds)
            if m:
                a, b2, y = int(m.group(1)), int(m.group(2)), m.group(3)
                data = f"{y}-{str(a).zfill(2)}-{str(b2).zfill(2)}"
            else:
                m2 = re.match(r"(\d{4})[\/\-](\d{2})[\/\-](\d{2})", ds)
                if m2:
                    data = ds[:10]

        nota_fis = norm_nf(nf_raw.split("/")[0]) if nf_raw and nf_raw != "nan" else ""
        regs.append({"data": data, "nome": nome, "valor": valor, "notaFis": nota_fis})

    return pd.DataFrame(regs)


# ─── Cruzamento ───────────────────────────────────────────────────────────────

def cruzar(dda: pd.DataFrame, sis: pd.DataFrame, limiar: float = LIMIAR_NOME) -> list[dict]:
    results = []
    used_sis = set()

    for _, d in dda.iterrows():
        best_sim = 0.0
        best_idx = None

        for idx, s in sis.iterrows():
            if idx in used_sis:
                continue
            sim = similaridade(d["beneficiario"], s["nome"])
            if sim < limiar:
                continue
            tol = tol_valor(max(d["valor"], s["valor"]))
            if abs(d["valor"] - s["valor"]) > tol:
                continue
            if sim > best_sim:
                best_sim = sim
                best_idx = idx

        if best_idx is not None:
            used_sis.add(best_idx)
            s = sis.loc[best_idx]
            ok_n  = best_sim >= LIMIAR_NOME
            ok_v  = abs(d["valor"] - s["valor"]) <= tol_valor(max(d["valor"], s["valor"]))
            ok_nf = bool(d["seuNum"] and s["notaFis"] and d["seuNum"] == s["notaFis"])
            results.append({
                "status": "ok",
                "dda_data": d["data"], "dda_benef": d["beneficiario"],
                "dda_valor": d["valor"], "dda_seuNum": d["seuNum"],
                "sys_data": s["data"], "sys_nome": s["nome"],
                "sys_valor": s["valor"], "sys_nf": s["notaFis"],
                "sim": best_sim, "okN": ok_n, "okV": ok_v, "okNF": ok_nf,
                "diff": d["valor"] - s["valor"],
            })
        else:
            results.append({
                "status": "soDda",
                "dda_data": d["data"], "dda_benef": d["beneficiario"],
                "dda_valor": d["valor"], "dda_seuNum": d["seuNum"],
                "sys_data": None, "sys_nome": None, "sys_valor": None, "sys_nf": None,
                "sim": 0.0, "okN": False, "okV": False, "okNF": False, "diff": None,
            })

    for idx, s in sis.iterrows():
        if idx not in used_sis:
            results.append({
                "status": "soSis",
                "dda_data": None, "dda_benef": None, "dda_valor": None, "dda_seuNum": None,
                "sys_data": s["data"], "sys_nome": s["nome"],
                "sys_valor": s["valor"], "sys_nf": s["notaFis"],
                "sim": 0.0, "okN": False, "okV": False, "okNF": False, "diff": None,
            })

    return results


# ─── Export XLSX ──────────────────────────────────────────────────────────────

def exportar_xlsx(results: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Conferência DDA"

    GREEN  = PatternFill("solid", fgColor="C6EFCE")
    RED    = PatternFill("solid", fgColor="FFC7CE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    BLUE_H = PatternFill("solid", fgColor="1F4E79")
    TEAL_H = PatternFill("solid", fgColor="1F6B75")
    GRAY_H = PatternFill("solid", fgColor="595959")
    WH     = Font(color="FFFFFF", bold=True)
    CENTER = Alignment(horizontal="center", vertical="center")
    thin   = Side(style="thin", color="CCCCCC")
    BDR    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Linha 1: grupos
    ws.merge_cells("A1:D1"); ws["A1"] = "📄 DDA BANCÁRIO"
    ws.merge_cells("E1:H1"); ws["E1"] = "📊 SISTEMA INTERNO"
    ws.merge_cells("I1:N1"); ws["I1"] = "RESULTADO"
    for cell, fill in [("A1", BLUE_H), ("E1", TEAL_H), ("I1", GRAY_H)]:
        ws[cell].fill = fill; ws[cell].font = WH; ws[cell].alignment = CENTER

    # Linha 2: colunas
    headers = [
        "Data Venc.", "Beneficiário Original", "Valor (R$)", "Seu Número",
        "Vencimento", "Terceiro", "Vlr. Nom.", "Nota Fis.",
        "Similaridade", "Nome?", "Valor?", "NF?", "Diferença", "Status",
    ]
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.alignment = CENTER
        cell.border = BDR

    SL = {"ok": "✅ Conferido", "soSis": "⚠️ Só no Sistema", "soDda": "🔴 Só no DDA"}
    for r in results:
        row = [
            fmt_data(r["dda_data"]), r["dda_benef"] or "", fmt_brl(r["dda_valor"]), r["dda_seuNum"] or "",
            fmt_data(r["sys_data"]), r["sys_nome"] or "", fmt_brl(r["sys_valor"]), r["sys_nf"] or "",
            f"{r['sim']*100:.0f}%" if r["sim"] else "",
            ("✅" if r["okN"] else "❌") if r["status"] == "ok" else "",
            ("✅" if r["okV"] else "❌") if r["status"] == "ok" else "",
            ("✅" if r["okNF"] else "❌") if r["status"] == "ok" else "",
            fmt_brl(r["diff"]) if r["diff"] is not None else "",
            SL[r["status"]],
        ]
        ws.append(row)
        row_num = ws.max_row
        fill = GREEN if r["status"] == "ok" else (RED if r["status"] == "soDda" else YELLOW)
        for col in range(1, 15):
            cell = ws.cell(row_num, col)
            cell.fill = fill
            cell.border = BDR
            cell.alignment = Alignment(vertical="center")

    # Larguras
    widths = [13, 38, 14, 16, 13, 38, 14, 16, 12, 8, 8, 8, 14, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Interface ────────────────────────────────────────────────────────────────

st.title("🔍 Conferência DDA Bancário × Sistema Interno")
st.caption("Detecção de boletos no DDA que não constam no sistema — anti-fraude")

# Sidebar de configurações
with st.sidebar:
    st.header("⚙️ Configurações")
    limiar_pct = st.slider(
        "Similaridade mínima de nome (%)",
        min_value=50, max_value=100, value=75, step=5,
        help="Quanto os nomes precisam ser parecidos para casar. Baixe se estiver perdendo registros."
    )
    limiar = limiar_pct / 100.0
    st.caption(f"Tolerância de valor: ±1% ou ±R$1,00 (o maior)")

col1, col2 = st.columns(2)

with col1:
    st.subheader("📄 DDA Bancário")
    file_dda = st.file_uploader(
        "Arraste ou selecione o arquivo DDA",
        type=["pdf", "xlsx", "xls", "csv"],
        key="dda",
    )
    st.caption("Colunas: Data Venc. · Beneficiário Original · Valor (R$) · Seu Número")

with col2:
    st.subheader("📊 Sistema Interno")
    file_sis = st.file_uploader(
        "Arraste ou selecione o relatório do sistema",
        type=["xlsx", "xls", "xlsm"],
        key="sis",
    )
    st.caption("Colunas: Vencimento · Terceiro · Vlr. Nom. · Nota Fis.")

st.divider()

if file_dda and file_sis:
    if st.button("🔍 Conferir", type="primary", use_container_width=True):
        with st.spinner("Lendo arquivos e cruzando dados..."):
            # Lê DDA
            dda_info = {}
            if file_dda.name.lower().endswith(".pdf"):
                dda = ler_dda_pdf(file_dda)
            else:
                dda, dda_info = ler_dda_excel(file_dda)

            # Lê Sistema
            sis = ler_sistema(file_sis)

        if dda.empty:
            st.error("❌ Nenhum registro extraído do DDA. Verifique o arquivo.")
            if dda_info.get("colunas_brutas"):
                st.info(f"Colunas encontradas no DDA: {dda_info['colunas_brutas']}")
            st.stop()
        if sis.empty:
            st.error("❌ Nenhum registro extraído do Sistema. Verifique o arquivo.")
            st.stop()

        # ── Diagnóstico ──
        with st.expander("🔎 Diagnóstico — dados lidos (clique para ver)", expanded=True):
            if dda_info.get("aviso"):
                st.warning(dda_info["aviso"])
            cols_info = dda_info.get("colunas_brutas", [])
            ci_d = dda_info.get("col_data", -1)
            bi_d = dda_info.get("col_benef", -1)
            vi_d = dda_info.get("col_valor", -1)
            si_d = dda_info.get("col_snum", -1)
            if cols_info:
                st.caption(f"**Cabeçalhos DDA:** {cols_info}")
                st.caption(
                    f"Colunas detectadas → Data: col {ci_d} `{cols_info[ci_d] if 0<=ci_d<len(cols_info) else '?'}`  |  "
                    f"Beneficiário: col {bi_d} `{cols_info[bi_d] if 0<=bi_d<len(cols_info) else '?'}`  |  "
                    f"Valor: col {vi_d} `{cols_info[vi_d] if 0<=vi_d<len(cols_info) else '?'}`  |  "
                    f"Seu Nº: col {si_d} `{cols_info[si_d] if 0<=si_d<len(cols_info) else 'não encontrado'}`"
                )
            dc1, dc2 = st.columns(2)
            with dc1:
                st.caption("**DDA — beneficiário e valor lidos:**")
                st.dataframe(dda[["beneficiario", "valor"]].head(15), use_container_width=True)
            with dc2:
                st.caption("**Sistema — terceiro e valor lidos:**")
                st.dataframe(sis[["nome", "valor"]].head(15), use_container_width=True)

        st.success(f"DDA: {len(dda)} registro(s)   |   Sistema: {len(sis)} registro(s)   |   Limiar: {limiar_pct}%")

        with st.spinner("Cruzando dados..."):
            results = cruzar(dda, sis, limiar=limiar)

        # ── Resumo ──
        n_ok    = sum(1 for r in results if r["status"] == "ok")
        n_soSis = sum(1 for r in results if r["status"] == "soSis")
        n_soDda = sum(1 for r in results if r["status"] == "soDda")

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total",          len(results))
        m2.metric("✅ Conferidos",   n_ok)
        m3.metric("⚠️ Só no Sistema", n_soSis)
        m4.metric("🔴 Só no DDA",    n_soDda)

        if n_soDda > 0:
            st.warning(f"⚠️ {n_soDda} boleto(s) no DDA sem correspondência no sistema — verifique possível fraude!")

        # ── Filtros ──
        filtro = st.radio(
            "Filtrar:",
            ["Todos", "✅ Conferidos", "⚠️ Só no Sistema", "🔴 Só no DDA"],
            horizontal=True,
        )
        mapa = {"Todos": None, "✅ Conferidos": "ok", "⚠️ Só no Sistema": "soSis", "🔴 Só no DDA": "soDda"}
        filtrado = [r for r in results if mapa[filtro] is None or r["status"] == mapa[filtro]]

        # ── Tabela ──
        SL = {"ok": "✅ Conferido", "soSis": "⚠️ Só no Sistema", "soDda": "🔴 Só no DDA"}
        rows_display = []
        for r in filtrado:
            sim_str = f"{r['sim']*100:.0f}%" if r["sim"] else "—"
            rows_display.append({
                "DDA Data":       fmt_data(r["dda_data"]),
                "DDA Beneficiário": r["dda_benef"] or "—",
                "DDA Valor":      fmt_brl(r["dda_valor"]),
                "DDA Seu Nº":     r["dda_seuNum"] or "—",
                "Sys Vencimento": fmt_data(r["sys_data"]),
                "Sys Terceiro":   r["sys_nome"] or "—",
                "Sys Vlr. Nom.":  fmt_brl(r["sys_valor"]),
                "Sys Nota Fis.":  r["sys_nf"] or "—",
                "Similaridade":   sim_str,
                "Nome ✓":         ("✅" if r["okN"] else "❌") if r["status"] == "ok" else "—",
                "Valor ✓":        ("✅" if r["okV"] else "❌") if r["status"] == "ok" else "—",
                "NF ✓":           ("✅" if r["okNF"] else "❌") if r["status"] == "ok" else "—",
                "Diferença":      fmt_brl(r["diff"]) if r["diff"] is not None else "—",
                "Status":         SL[r["status"]],
            })

        df_display = pd.DataFrame(rows_display)

        def color_row(row):
            status = row["Status"]
            if "Conferido" in status:
                return ["background-color: #C6EFCE"] * len(row)
            elif "Sistema" in status:
                return ["background-color: #FFEB9C"] * len(row)
            else:
                return ["background-color: #FFC7CE"] * len(row)

        st.dataframe(
            df_display.style.apply(color_row, axis=1),
            use_container_width=True,
            height=min(600, 60 + len(rows_display) * 35),
        )

        # ── Export ──
        xlsx_bytes = exportar_xlsx(results)
        today = date.today().strftime("%Y-%m-%d")
        st.download_button(
            "⬇️ Exportar resultado (.xlsx)",
            data=xlsx_bytes,
            file_name=f"conferencia_dda_{today}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
else:
    st.info("👆 Faça o upload dos dois arquivos para habilitar a conferência.")
